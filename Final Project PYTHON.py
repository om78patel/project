import openpyxl
from datetime import datetime
import os.path

class Employee:
    def _init_(self):
        self.name = ""
        self.hours_worked = 0
        self.hourly_rate = 0

def calculate_salary(employee):
    return (employee.hours_worked * employee.hourly_rate) / 1.13

def format_employee_data(employees):
    formatted_data = "Employee Report:\n\n"
    total_salary = 0
    Tax = 0

    for employee in employees:
        salary = calculate_salary(employee)
        Tax = (employee.hours_worked * employee.hourly_rate) - salary
        formatted_data += f"Name: {employee.name}\n"
        formatted_data += f"Hours Worked: {employee.hours_worked}\n"
        formatted_data += f"Hourly Rate: ${employee.hourly_rate:.2f}\n"
        formatted_data += f"Taxes: ${Tax:.2f}\n"
        formatted_data += f"Salary: ${salary:.2f}\n\n"
        total_salary += salary  # Accumulate total salary

    formatted_data += f"Total Salary for all employees: ${total_salary:.2f}\n"
    return formatted_data

def create_excel_report(employees):
    filename = "Employee_Report3.xlsx"
    
    # Check if the file exists, if not, create a new one
    if not os.path.isfile(filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        # Add headers
        headers = ["Name", "Hours Worked", "Hourly Rate", "Taxes", "Salary", "Date and Time"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
    else:
        # Open the existing workbook
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

    # Add employee data
    for row_num, employee in enumerate(employees, ws.max_row + 1):
        ws.cell(row=row_num, column=1, value=employee.name)
        ws.cell(row=row_num, column=2, value=employee.hours_worked)
        ws.cell(row=row_num, column=3, value=employee.hourly_rate)
        ws.cell(row=row_num, column=4, value=(employee.hours_worked * employee.hourly_rate) - calculate_salary(employee))
        ws.cell(row=row_num, column=5, value=calculate_salary(employee))
        ws.cell(row=row_num, column=6, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    # Add total salary
    ws.cell(row=ws.max_row + 1, column=1, value="Total Salary")
    ws.cell(row=ws.max_row, column=5, value=sum(calculate_salary(emp) for emp in employees))

    try:
        # Save the workbook
        wb.save(filename)
        print(f"Excel report updated successfully at {filename}")
    except PermissionError:
        print(f"PermissionError: Unable to save the file. Check write permissions for the specified path.")

def create_graph(employees):
    import matplotlib.pyplot as plt
    names = [employee.name for employee in employees]
    salaries = [calculate_salary(employee) for employee in employees]
    plt.bar(names, salaries)
    plt.xlabel("Employee Names")
    plt.ylabel("Salary")
    plt.title("Employee Salaries")
    plt.show()

def get_valid_input(prompt, data_type=float):
    while True:
        try:
            user_input = data_type(input(prompt))
            return user_input
        except ValueError:
            print("Invalid input. Please enter a valid number.")

# Example usage
num_employees = int(get_valid_input("Enter the number of employees: ", int))
employees = []
for _ in range(num_employees):
    employee = Employee()
    employee.name = input("Enter employee name: ")
    employee.hours_worked = get_valid_input("Enter hours worked: ")
    employee.hourly_rate = get_valid_input("Enter hourly rate: ")
    employees.append(employee)
formatted_report = format_employee_data(employees)
print(formatted_report)

# Create Excel report and update the existing file
create_excel_report(employees)

# Create graph
create_graph(employees)
